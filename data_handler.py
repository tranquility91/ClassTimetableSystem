# data_handler.py
import csv
import os
from datetime import datetime
import openpyxl

def save_to_csv(solution):
    """將分配結果保存為CSV文件"""
    # 指定儲存路徑
    save_path = r"C:\Users\USER\OneDrive\桌面"
    
    # 生成檔案名稱，包含時間戳
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"course_allocation_{timestamp}.csv"
    
    # 組合完整的檔案路徑
    full_path = os.path.join(save_path, filename)
    
    # 確保目標資料夾存在
    os.makedirs(save_path, exist_ok=True)
    
    # 準備CSV寫入
    with open(full_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        # 寫入標題行
        writer.writerow(['老師', '年級', '課程名稱', '組別', '節數', '人次'])
        
        # 寫入每位老師的課程資料
        for teacher, load in solution.items():
            for course in sorted(load['課程'], 
                              key=lambda x: (x['課程名稱'], x['年級'], x['組別'])):
                writer.writerow([
                    teacher,
                    course['年級'],
                    course['課程名稱'],
                    course['組別'],
                    course['節數'],
                    course['人次']
                ])
            # 在每位老師的課程後加入合計行
            writer.writerow([
                f"{teacher}合計",
                '',
                '',
                '',
                load['節數'],
                load['人次']
            ])
            # 加入空行分隔
            writer.writerow([])
    
    print(f"\n分配結果已保存至 {full_path}")

def update_excel_with_solution(solution, excel_path):
    """將分配結果更新到Excel"""
    wb = openpyxl.load_workbook(excel_path)
    
    # 建立新工作表
    result_sheet = wb.create_sheet(f"配課結果_{datetime.now().strftime('%m%d_%H%M')}")
    
    # 設定標題行
    headers = ['老師', '年級', '課程名稱', '組別', '節數', '人數', '人次', 
              '備課組別', '總節數', '總人次']
    for col, header in enumerate(headers, 1):
        result_sheet.cell(row=1, column=col, value=header)
    
    # 寫入課程資料
    current_row = 2
    for teacher, load in solution.items():
        # 計算備課組別
        special_groups = len([c for c in load["課程"] 
                            if c["課程名稱"] in ["社會", "動作", "學習"]])
        chinese_math_groups = len(set(f"{c['課程名稱']}_{c['年級']}" 
                                    for c in load["課程"] 
                                    if c["課程名稱"] in ["國語", "數學"]))
        
        # 寫入每門課程
        for course in sorted(load['課程'], 
                           key=lambda x: (x['課程名稱'], x['年級'], x['組別'])):
            result_sheet.cell(row=current_row, column=1, value=teacher)
            result_sheet.cell(row=current_row, column=2, value=course['年級'])
            result_sheet.cell(row=current_row, column=3, value=course['課程名稱'])
            result_sheet.cell(row=current_row, column=4, value=course['組別'])
            result_sheet.cell(row=current_row, column=5, value=course['節數'])
            result_sheet.cell(row=current_row, column=6, value=course['人數'])
            result_sheet.cell(row=current_row, column=7, value=course['人次'])
            current_row += 1
        
        # 寫入合計行
        result_sheet.cell(row=current_row, column=1, value=f"{teacher}合計")
        result_sheet.cell(row=current_row, column=8, 
                         value=f"特需{special_groups}組 + 國數{chinese_math_groups}組")
        result_sheet.cell(row=current_row, column=9, value=load['節數'])
        result_sheet.cell(row=current_row, column=10, value=load['人次'])
        
        current_row += 2  # 空一行
    
    # 儲存新檔案
    file_dir = os.path.dirname(excel_path)
    file_name = os.path.basename(excel_path)
    name, ext = os.path.splitext(file_name)
    new_file_path = os.path.join(file_dir, f"{name}_配課結果{ext}")
    
    wb.save(new_file_path)
    print(f"\n分配結果已更新至 {new_file_path}")